from decimal import Decimal
import hashlib
import hmac
from io import BytesIO
import json
import time
from unittest.mock import patch

from django.contrib.auth.models import Group
from django.core.cache import cache
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from django_otp.oath import totp
from django_otp.plugins.otp_totp.models import TOTPDevice
from rest_framework.test import APIClient

from users.models import User
from products.models import Product
from .models import Order, OrderItem, SeevCashWebhookEvent
from .promo_models import PromoCode

CREATE_URL = '/api/orders/create/'
VERIFY_URL = '/api/orders/verify-payment/'


def seevcash_ok(reference='SEEV-SESSION', amount=5000, currency='GHS', txn_id='txn-99'):
    return {
        'status': 'completed', 'reference': reference, 'currency': currency,
        'amount': amount, 'final_amount': amount, 'id': txn_id,
    }


class PaymentVerificationTests(TestCase):
    def setUp(self):
        cache.clear()
        self.user = User.objects.create_user(
            email='c@example.com', password='x', first_name='C', last_name='U',
            email_verified=True,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        self.order = Order.objects.create(
            user=self.user, reference='LO-PAYTEST', delivery_address='Accra',
            total_amount=Decimal('50.00'), discount_amount=Decimal('0.00'),
            payment_status='pending', status='pending', order_source='seevcash',
            payment_provider='seevcash', checkout_reference='SEEV-SESSION',
        )

    @override_settings(SEEVCASH_SECRET_KEY='sandbox-key')
    @patch('orders.views.create_checkout')
    def test_checkout_initialization_uses_stable_order_idempotency_key(self, mock_create):
        from legitorganic.seevcash import CheckoutSession
        mock_create.return_value = CheckoutSession(
            reference='SEEV-NEW', checkout_url='https://pay.seevplus.com/SEEV-NEW',
            status='pending',
        )
        response = self.client.post('/api/orders/LO-PAYTEST/checkout/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['checkout_url'], 'https://pay.seevplus.com/SEEV-NEW')
        self.assertEqual(mock_create.call_args.kwargs['idempotency_key'], f'order-{self.order.pk}-checkout')
        self.order.refresh_from_db()
        self.assertEqual(self.order.checkout_reference, 'SEEV-NEW')

    def test_fails_closed_without_checkout_reference(self):
        self.order.checkout_reference = ''
        self.order.save(update_fields=['checkout_reference'])
        resp = self.client.post(VERIFY_URL, {'order_reference': 'LO-PAYTEST'}, format='json')
        self.assertEqual(resp.status_code, 400)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'pending')

    @patch('orders.views.verify_checkout')
    def test_fails_closed_when_seevcash_times_out(self, mock_verify):
        from legitorganic.seevcash import SeevCashError
        mock_verify.side_effect = SeevCashError('timeout')
        resp = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(resp.status_code, 502)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'pending')

    @patch('orders.views.verify_checkout')
    def test_declined_payment_marked_failed(self, mock_verify):
        mock_verify.return_value = {'status': 'failed'}
        resp = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(resp.status_code, 402)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'failed')

    @patch('orders.views.verify_checkout')
    def test_amount_underpayment_rejected(self, mock_verify):
        # Order is GHS 50.00 == 5000 pesewas; provider reports only 1000.
        mock_verify.return_value = seevcash_ok(amount=1000)
        resp = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(resp.status_code, 400)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'pending')

    @patch('orders.views.verify_checkout')
    def test_currency_mismatch_rejected(self, mock_verify):
        mock_verify.return_value = seevcash_ok(currency='NGN')
        resp = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(resp.status_code, 400)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'pending')

    @patch('orders.views.verify_checkout')
    def test_reference_mismatch_rejected(self, mock_verify):
        mock_verify.return_value = seevcash_ok(reference='SOMETHING-ELSE')
        resp = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(resp.status_code, 400)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'pending')

    @patch('users.sms.send_order_status_sms')
    @patch('users.emails.resend.Emails.send')
    @patch('orders.views.verify_checkout')
    def test_valid_payment_marks_paid(self, mock_verify, mock_email, mock_sms):
        mock_verify.return_value = seevcash_ok(amount=5000, txn_id='txn-12345')
        resp = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(resp.status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'success')
        self.assertEqual(self.order.status, 'processing')
        self.assertEqual(self.order.provider_transaction_id, 'txn-12345')

    @patch('users.sms.send_order_status_sms')
    @patch('users.emails.resend.Emails.send')
    @patch('orders.views.verify_checkout')
    def test_repeated_verification_transitions_exactly_once(self, mock_verify, mock_email, mock_sms):
        # Guards the transition side effects (confirmation/status emails, commissions)
        # against firing twice. True parallelism needs Postgres + threads; here we
        # assert the idempotent recheck: a second completion does no extra work.
        mock_verify.return_value = seevcash_ok(amount=5000, txn_id='txn-777')
        r1 = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(r1.status_code, 200)
        emails_after_first = mock_email.call_count
        self.assertGreaterEqual(emails_after_first, 1)

        r2 = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(r2.status_code, 200)
        self.assertEqual(mock_email.call_count, emails_after_first)  # no duplicate side effects
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'success')

    @patch('orders.views.verify_checkout')
    def test_already_paid_is_idempotent(self, mock_verify):
        self.order.payment_status = 'success'
        self.order.status = 'processing'
        self.order.save(update_fields=['payment_status', 'status'])
        resp = self.client.post(VERIFY_URL, {'reference': 'SEEV-SESSION'}, format='json')
        self.assertEqual(resp.status_code, 200)
        mock_verify.assert_not_called()  # no re-verification / re-processing


@override_settings(SEEVCASH_WEBHOOK_SECRET='whsec_test')
class SeevCashWebhookTests(TestCase):
    url = '/api/orders/seevcash/webhook/'

    def setUp(self):
        self.client = APIClient()
        self.order = Order.objects.create(
            reference='LO-WEBHOOK', delivery_address='Accra',
            total_amount=Decimal('50.00'), payment_status='pending',
            status='pending', order_source='seevcash', payment_provider='seevcash',
            checkout_reference='SEEV-WEBHOOK',
        )

    def post_event(self, event_id='evt-1', event_type='payment.succeeded', timestamp=None,
                   signature_secret='whsec_test', payload=None):
        payload = payload or {'data': {'reference': 'SEEV-WEBHOOK'}}
        raw = json.dumps(payload, separators=(',', ':')).encode()
        timestamp = str(timestamp if timestamp is not None else int(time.time()))
        digest = hmac.new(
            signature_secret.encode(), timestamp.encode() + b'.' + raw, hashlib.sha256
        ).hexdigest()
        return self.client.generic(
            'POST', self.url, raw, content_type='application/json',
            HTTP_X_SEEV_EVENT_ID=event_id,
            HTTP_X_SEEV_EVENT_TYPE=event_type,
            HTTP_X_SEEV_TIMESTAMP=timestamp,
            HTTP_X_SEEV_SIGNATURE=f'v1={digest}',
        )

    def test_invalid_signature_is_rejected(self):
        response = self.post_event(signature_secret='wrong')
        self.assertEqual(response.status_code, 401)
        self.assertFalse(SeevCashWebhookEvent.objects.exists())

    def test_stale_timestamp_is_rejected(self):
        response = self.post_event(timestamp=int(time.time()) - 301)
        self.assertEqual(response.status_code, 401)

    @patch('orders.views.verify_checkout')
    def test_signed_success_is_verified_and_completed(self, mock_verify):
        mock_verify.return_value = seevcash_ok(reference='SEEV-WEBHOOK')
        response = self.post_event()
        self.assertEqual(response.status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'success')
        self.assertEqual(SeevCashWebhookEvent.objects.get().status, 'processed')

    def test_signed_success_accepts_nested_checkout_reference(self):
        self.order.checkout_reference = 'PAY-SEEV-WEBHOOK'
        self.order.save(update_fields=['checkout_reference'])
        response = self.post_event(payload={
            'data': {
                'transaction': {
                    'checkoutReference': 'PAY-SEEV-WEBHOOK',
                    'amount': 5000,
                    'currency': 'GHS',
                    'paymentId': 'txn-nested',
                }
            }
        })
        self.assertEqual(response.status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'success')
        self.assertEqual(self.order.provider_transaction_id, 'txn-nested')

    @patch('orders.views.verify_checkout')
    def test_same_delivery_and_retry_delivery_are_idempotent(self, mock_verify):
        mock_verify.return_value = seevcash_ok(reference='SEEV-WEBHOOK')
        self.assertEqual(self.post_event(event_id='evt-first').status_code, 200)
        first_verifications = mock_verify.call_count
        self.assertEqual(self.post_event(event_id='evt-first').status_code, 200)
        self.assertEqual(mock_verify.call_count, first_verifications)
        self.assertEqual(self.post_event(event_id='evt-retry').status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'success')
        self.assertEqual(SeevCashWebhookEvent.objects.count(), 2)

    def test_signed_failure_marks_only_pending_order_failed(self):
        response = self.post_event(event_type='payment.failed')
        self.assertEqual(response.status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'failed')


class ReceiptAuthTests(TestCase):
    def setUp(self):
        cache.clear()
        self.owner = User.objects.create_user(
            email='own@example.com', password='x', first_name='O', last_name='W',
            email_verified=True,
        )
        self.other = User.objects.create_user(
            email='oth@example.com', password='x', first_name='O', last_name='T',
            email_verified=True,
        )
        self.order = Order.objects.create(
            user=self.owner, reference='LO-RCPT1', delivery_address='Accra',
            total_amount=Decimal('20.00'), payment_status='success',
            status='processing', order_source='paystack',
        )

    def test_anonymous_cannot_download_receipt(self):
        resp = APIClient().get('/api/orders/LO-RCPT1/receipt/')
        self.assertIn(resp.status_code, (401, 403))

    def test_non_owner_gets_404(self):
        c = APIClient()
        c.force_authenticate(self.other)
        resp = c.get('/api/orders/LO-RCPT1/receipt/')
        self.assertEqual(resp.status_code, 404)

    def test_owner_can_download_receipt(self):
        c = APIClient()
        c.force_authenticate(self.owner)
        resp = c.get('/api/orders/LO-RCPT1/receipt/')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')


class AtomicOrderCreationTests(TestCase):
    def setUp(self):
        cache.clear()
        self.client = APIClient()
        self.product = Product.objects.create(name='Tomatoes', price=Decimal('10.00'))

    def _payload(self, items, **extra):
        base = {
            'items': items,
            'delivery_address': '10 Farm Road, Accra, Greater Accra',
            'phone_number': '0200000000',
            'street_address': 'Farm Road',
            'house_number': '10',
            'city': 'Accra',
            'delivery_region': 'Greater Accra',
            'order_source': 'whatsapp',
            'guest_name': 'Guest',
            'guest_phone': '0200000000',
            'guest_email': 'guest@example.com',
        }
        base.update(extra)
        return base

    def test_guest_order_requires_phone_and_structured_address(self):
        payload = self._payload([{'product_id': self.product.id, 'quantity': 1}])
        payload['guest_phone'] = ''
        payload['phone_number'] = ''
        payload['street_address'] = ''
        resp = self.client.post(CREATE_URL, payload, format='json')
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(Order.objects.count(), 0)

    def test_authenticated_checkout_updates_customer_delivery_profile(self):
        user = User.objects.create_user(
            email='buyer@example.com', password='x', first_name='Ama', last_name='Mensah',
            email_verified=True,
        )
        self.client.force_authenticate(user)
        payload = self._payload([{'product_id': self.product.id, 'quantity': 1}])
        resp = self.client.post(CREATE_URL, payload, format='json')
        self.assertEqual(resp.status_code, 201)
        user.refresh_from_db()
        order = Order.objects.get(reference=resp.data['reference'])
        self.assertEqual(user.phone_number, '0200000000')
        self.assertEqual(user.street_address, 'Farm Road')
        self.assertEqual(order.guest_phone, '0200000000')
        self.assertEqual(order.delivery_address, '10, Farm Road, Accra, Greater Accra')

    def test_invalid_product_leaves_no_orphan_order(self):
        before = Order.objects.count()
        resp = self.client.post(CREATE_URL, self._payload([
            {'product_id': self.product.id, 'quantity': 2},
            {'product_id': 999999, 'quantity': 1},
        ]), format='json')
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(Order.objects.count(), before)
        self.assertEqual(OrderItem.objects.count(), 0)

    def test_valid_order_created_with_correct_total(self):
        resp = self.client.post(CREATE_URL, self._payload([
            {'product_id': self.product.id, 'quantity': 3},
        ]), format='json')
        self.assertEqual(resp.status_code, 201)
        order = Order.objects.get(reference=resp.data['reference'])
        self.assertEqual(order.total_amount, Decimal('30.00'))
        self.assertEqual(order.items.count(), 1)

    def test_valid_promo_increments_usage_once(self):
        promo = PromoCode.objects.create(
            code='SAVE10', discount_type=PromoCode.DISCOUNT_FIXED,
            discount_value=Decimal('10.00'), is_active=True,
        )
        resp = self.client.post(CREATE_URL, self._payload([
            {'product_id': self.product.id, 'quantity': 2},
        ], promo_code='SAVE10'), format='json')
        self.assertEqual(resp.status_code, 201)
        promo.refresh_from_db()
        self.assertEqual(promo.times_used, 1)
        order = Order.objects.get(reference=resp.data['reference'])
        self.assertEqual(order.discount_amount, Decimal('10.00'))


class OwnerOrderReportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            email='report-buyer@example.com', password='x', first_name='Esi',
            last_name='Owusu', phone_number='0244000000', email_verified=True,
        )
        self.order = Order.objects.create(
            user=self.user, reference='LO-REPORT1',
            delivery_address='10 Farm Road, Accra, Greater Accra',
            guest_phone='0244000000', total_amount=Decimal('25.00'),
            status='pending', payment_status='pending', order_source='paystack',
        )

    @override_settings(ORDER_REPORT_EMAIL='legitorganic9@gmail.com')
    @patch('users.emails.resend.Emails.send')
    def test_payment_and_delivery_reports_are_each_sent_once(self, mock_send):
        with self.captureOnCommitCallbacks(execute=True):
            self.order.payment_status = 'success'
            self.order.status = 'processing'
            self.order.save(update_fields=['payment_status', 'status'])
        self.order.refresh_from_db()
        self.assertIsNotNone(self.order.payment_report_sent_at)
        self.assertEqual(mock_send.call_count, 2)  # customer status + owner report
        owner_payload = mock_send.call_args_list[-1].args[0]
        self.assertEqual(owner_payload['to'], ['legitorganic9@gmail.com'])
        self.assertIn('Payment received', owner_payload['subject'])

        # Saving the same successful state again must not send another report.
        with self.captureOnCommitCallbacks(execute=True):
            self.order.save(update_fields=['updated_at'])
        self.assertEqual(mock_send.call_count, 2)

        with self.captureOnCommitCallbacks(execute=True):
            self.order.status = 'delivered'
            self.order.save(update_fields=['status'])
        self.order.refresh_from_db()
        self.assertIsNotNone(self.order.delivery_report_sent_at)
        self.assertEqual(mock_send.call_count, 4)  # customer status + owner report
        self.assertIn('Order delivered', mock_send.call_args_list[-1].args[0]['subject'])

    @patch('users.emails.resend.Emails.send', side_effect=RuntimeError('mail down'))
    def test_failed_report_remains_retryable(self, _mock_send):
        with self.captureOnCommitCallbacks(execute=True):
            self.order.payment_status = 'success'
            self.order.save(update_fields=['payment_status'])
        self.order.refresh_from_db()
        self.assertIsNone(self.order.payment_report_sent_at)
        self.assertEqual(self.order.payment_report_attempts, 1)

        with patch('users.emails.resend.Emails.send') as retry_send:
            call_command('retry_order_reports')
        self.order.refresh_from_db()
        self.assertIsNotNone(self.order.payment_report_sent_at)
        self.assertEqual(self.order.payment_report_attempts, 2)
        retry_send.assert_called_once()


class OrderExportAdminTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_superuser(
            email='owner@legitorganic.com', password='x', first_name='Owner', last_name='User'
        )
        self.client.force_login(self.staff)
        self.match = Order.objects.create(
            reference='LO-EXPORT-MATCH', delivery_address='Accra', guest_name='Ama Match',
            guest_phone='0244000000', total_amount=Decimal('30.00'),
            status='delivered', payment_status='success', order_source='whatsapp',
        )
        Order.objects.create(
            reference='LO-EXPORT-OTHER', delivery_address='Kumasi', guest_name='Kojo Other',
            guest_phone='0200000000', total_amount=Decimal('15.00'),
            status='pending', payment_status='pending', order_source='paystack',
        )

    def test_orders_page_shows_visible_export_workspace(self):
        response = self.client.get(reverse('admin:orders_order_changelist'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Order reporting')
        self.assertContains(response, 'Download Excel')

    def test_export_filters_customer_status_payment_and_channel(self):
        response = self.client.get(reverse('admin:orders-export-all'), {
            'customer': 'Ama Match', 'status': 'delivered',
            'payment_status': 'success', 'source': 'whatsapp',
        })
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook['Orders Summary']
        references = [sheet.cell(row=row, column=1).value for row in range(5, sheet.max_row)]
        self.assertIn('LO-EXPORT-MATCH', references)
        self.assertNotIn('LO-EXPORT-OTHER', references)
        self.assertIn('Customer/reference: Ama Match', sheet['A2'].value)

    def test_invalid_or_reversed_dates_return_to_orders_with_message(self):
        url = reverse('admin:orders-export-all')
        response = self.client.get(url, {'date_from': 'not-a-date'})
        self.assertRedirects(response, reverse('admin:orders_order_changelist'))


class OrderAdminSecurityTests(TestCase):
    def setUp(self):
        call_command('setup_groups', verbosity=0)
        self.owner = User.objects.create_superuser(
            email='secure-owner@example.com', password='StrongPass123!',
            first_name='Secure', last_name='Owner',
        )
        self.device = TOTPDevice.objects.create(
            user=self.owner, name='Authenticator', confirmed=True,
        )
        self.order = Order.objects.create(
            reference='LO-MANUAL-SECURE', delivery_address='Accra',
            guest_name='Customer', guest_phone='0244000000',
            total_amount=Decimal('40.00'), status='pending',
            payment_status='pending', order_source='whatsapp',
        )

    def _token(self, device=None):
        device = device or self.device
        return str(totp(
            device.bin_key, step=device.step, t0=device.t0,
            digits=device.digits, drift=device.drift,
        )).zfill(device.digits)

    @staticmethod
    def _inline_management():
        return {
            'items-TOTAL_FORMS': '0',
            'items-INITIAL_FORMS': '0',
            'items-MIN_NUM_FORMS': '0',
            'items-MAX_NUM_FORMS': '1000',
        }

    @patch('users.emails.resend.Emails.send')
    def test_owner_manual_payment_correction_requires_reason_password_and_2fa(self, _mail):
        self.client.force_login(self.owner)
        url = reverse('admin:orders_order_change', args=[self.order.pk])
        denied_payload = {
            'status': 'pending', 'payment_status': 'success',
            'payment_change_reason': '', 'current_password': 'wrong', 'otp_token': '000000',
            '_save': 'Save',
            **self._inline_management(),
        }
        denied = self.client.post(url, denied_payload)
        self.assertEqual(denied.status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'pending')

        allowed = self.client.post(url, {
            'status': 'pending', 'payment_status': 'success',
            'payment_change_reason': 'MoMo receipt matched',
            'current_password': 'StrongPass123!', 'otp_token': self._token(),
            '_save': 'Save',
            **self._inline_management(),
        })
        self.assertEqual(allowed.status_code, 302)
        self.order.refresh_from_db()
        self.assertEqual(self.order.payment_status, 'success')
        from security.models import AuditEvent
        event = AuditEvent.objects.get(action='order.payment_corrected')
        self.assertEqual(event.reason, 'MoMo receipt matched')

    def test_operations_can_change_fulfilment_but_not_payment(self):
        operator = User.objects.create_user(
            email='ops@legitorganic.com', password='StrongPass123!', first_name='Op',
            last_name='User', is_staff=True, email_verified=True,
        )
        operator.groups.add(Group.objects.get(name='Operations'))
        self.client.force_login(operator)
        url = reverse('admin:orders_order_change', args=[self.order.pk])
        response = self.client.post(url, {
            'status': 'processing', 'payment_status': 'success', '_save': 'Save',
            **self._inline_management(),
        })
        self.assertEqual(response.status_code, 302)
        self.order.refresh_from_db()
        self.assertEqual(self.order.status, 'processing')
        self.assertEqual(self.order.payment_status, 'pending')
