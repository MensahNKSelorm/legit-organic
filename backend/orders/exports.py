import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from collections import Counter

# Brand colors
FOREST_GREEN = '0D3B2A'
GHANA_GOLD   = 'F4C430'
LIGHT_GREEN  = 'E8F5E9'
WHITE        = 'FFFFFF'
GREY         = 'F5F5F5'


def generate_orders_excel(orders, date_from=None, date_to=None, status_filter=None):
    wb = openpyxl.Workbook()

    # ── Shared styles ────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color=WHITE, size=11)
    header_fill  = PatternFill(fill_type='solid', fgColor=FOREST_GREEN)
    header_align = Alignment(horizontal='center', vertical='center')

    # ── Sheet 1: Orders Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Orders Summary'

    # Title row
    ws1.merge_cells('A1:L1')
    title_cell = ws1['A1']
    title_cell.value = 'LEGIT ORGANIC LIMITED — ORDER REPORT'
    title_cell.font = Font(bold=True, size=14, color=FOREST_GREEN)
    title_cell.alignment = Alignment(horizontal='center')
    ws1.row_dimensions[1].height = 30

    # Date / generated-at row
    ws1.merge_cells('A2:L2')
    date_cell = ws1['A2']
    generated_at = timezone.now().strftime('%d %B %Y at %H:%M')
    date_range = ''
    if date_from and date_to:
        date_range = f' | Period: {date_from} to {date_to}'
    elif date_from:
        date_range = f' | From: {date_from}'
    elif date_to:
        date_range = f' | To: {date_to}'
    date_cell.value = f'Generated: {generated_at}{date_range}'
    date_cell.font = Font(size=10, color='666666')
    date_cell.alignment = Alignment(horizontal='center')
    ws1.row_dimensions[2].height = 20

    # Empty spacer row
    ws1.row_dimensions[3].height = 10

    # Column headers
    headers = [
        'Reference', 'Date', 'Customer Name', 'Customer Email',
        'Phone', 'Order Source', 'Status', 'Payment Status',
        'Delivery Address', 'Promo Code', 'Discount (GH₵)', 'Total (GH₵)',
    ]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(bottom=Side(style='medium', color=GHANA_GOLD))
    ws1.row_dimensions[4].height = 25

    # Data rows
    total_revenue  = 0
    total_discount = 0

    for row_idx, order in enumerate(orders, 5):
        fill = PatternFill(
            fill_type='solid',
            fgColor=LIGHT_GREEN if row_idx % 2 == 0 else WHITE,
        )

        if order.user:
            customer_name  = f'{order.user.first_name} {order.user.last_name}'.strip()
            customer_email = order.user.email
            customer_phone = getattr(order.user, 'phone_number', None) or '-'
        else:
            customer_name  = order.guest_name  or 'Guest'
            customer_email = order.guest_email or '-'
            customer_phone = order.guest_phone or '-'

        discount = float(order.discount_amount or 0)
        total    = float(order.total_amount or 0)
        total_revenue  += total
        total_discount += discount

        promo = order.promo_code.code if order.promo_code_id else '-'

        row_data = [
            order.reference,
            order.created_at.strftime('%d/%m/%Y %H:%M'),
            customer_name,
            customer_email,
            customer_phone,
            order.get_order_source_display(),
            order.get_status_display(),
            order.get_payment_status_display(),
            order.delivery_address,
            promo,
            discount,
            total,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col in (11, 12):
                cell.number_format = '#,##0.00'

        ws1.row_dimensions[row_idx].height = 20

    # Totals row
    total_row = len(orders) + 5
    ws1.cell(row=total_row, column=10, value='TOTALS').font = Font(bold=True)

    disc_cell = ws1.cell(row=total_row, column=11, value=total_discount)
    disc_cell.font = Font(bold=True, color=FOREST_GREEN)
    disc_cell.number_format = '#,##0.00'

    rev_cell = ws1.cell(row=total_row, column=12, value=total_revenue)
    rev_cell.font = Font(bold=True, color=FOREST_GREEN, size=12)
    rev_cell.number_format = '#,##0.00'
    rev_cell.fill = PatternFill(fill_type='solid', fgColor=GHANA_GOLD)

    # Column widths
    col_widths = [18, 16, 20, 28, 16, 14, 18, 18, 35, 14, 14, 14]
    for col, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = width

    ws1.freeze_panes = 'A5'

    # ── Sheet 2: Order Items ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Order Items')

    item_headers = [
        'Order Reference', 'Date', 'Customer', 'Product Name',
        'Quantity', 'Unit Price (GH₵)', 'Subtotal (GH₵)',
    ]
    for col, header in enumerate(item_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws2.row_dimensions[1].height = 25

    item_row = 2
    for order in orders:
        if order.user:
            customer = f'{order.user.first_name} {order.user.last_name}'.strip()
        else:
            customer = order.guest_name or 'Guest'

        for item in order.items.all():
            product_name = item.product.name if item.product else 'Deleted Product'
            unit_price   = float(item.unit_price)
            subtotal     = unit_price * item.quantity

            row_data = [
                order.reference,
                order.created_at.strftime('%d/%m/%Y'),
                customer,
                product_name,
                item.quantity,
                unit_price,
                subtotal,
            ]

            fill = PatternFill(
                fill_type='solid',
                fgColor=GREY if item_row % 2 == 0 else WHITE,
            )
            for col, value in enumerate(row_data, 1):
                cell = ws2.cell(row=item_row, column=col, value=value)
                cell.fill = fill
                if col in (6, 7):
                    cell.number_format = '#,##0.00'
            item_row += 1

    item_col_widths = [18, 12, 22, 30, 10, 18, 16]
    for col, width in enumerate(item_col_widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = 'A2'

    # ── Sheet 3: Summary Stats ───────────────────────────────────────────────
    ws3 = wb.create_sheet('Summary Stats')

    stats_title = ws3.cell(row=1, column=1, value='ORDER STATISTICS')
    stats_title.font = Font(bold=True, size=14, color=FOREST_GREEN)

    status_counts = Counter(o.status       for o in orders)
    source_counts = Counter(o.order_source for o in orders)

    avg_order_value = round(total_revenue / len(orders), 2) if orders else 0

    stats = [
        ('',                             ''),
        ('OVERVIEW',                     ''),
        ('Total Orders',                 len(orders)),
        ('Total Revenue (GH₵)',          round(total_revenue, 2)),
        ('Total Discounts Given (GH₵)',  round(total_discount, 2)),
        ('Average Order Value (GH₵)',    avg_order_value),
        ('',                             ''),
        ('BY STATUS',                    ''),
        ('WhatsApp Pending',             status_counts.get('whatsapp_pending', 0)),
        ('Paid',                         status_counts.get('paid', 0)),
        ('Processing',                   status_counts.get('processing', 0)),
        ('Shipped',                      status_counts.get('shipped', 0)),
        ('Delivered',                    status_counts.get('delivered', 0)),
        ('Cancelled',                    status_counts.get('cancelled', 0)),
        ('',                             ''),
        ('BY SOURCE',                    ''),
        ('WhatsApp Orders',              source_counts.get('whatsapp', 0)),
        ('Paystack Orders',              source_counts.get('paystack', 0)),
    ]

    for row_idx, (label, value) in enumerate(stats, 2):
        label_cell = ws3.cell(row=row_idx, column=1, value=label)
        value_cell = ws3.cell(row=row_idx, column=2, value=value)

        if label in ('OVERVIEW', 'BY STATUS', 'BY SOURCE'):
            label_cell.font = Font(bold=True, color=WHITE)
            label_cell.fill = PatternFill(fill_type='solid', fgColor=FOREST_GREEN)
            ws3.merge_cells(f'A{row_idx}:B{row_idx}')
        elif label:
            value_cell.font = Font(bold=True, color=FOREST_GREEN)
            if 'GH₵' in label:
                value_cell.number_format = '#,##0.00'

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 20

    # ── Build filename and return response ───────────────────────────────────
    now      = timezone.now()
    date_str = now.strftime('%Y-%m-%d')
    time_str = now.strftime('%H%M')

    if date_from and date_to:
        period = f'{date_from}_to_{date_to}'
    elif date_from:
        period = f'from_{date_from}'
    elif date_to:
        period = f'to_{date_to}'
    else:
        period = 'AllTime'

    status_str = status_filter.replace(' ', '') if status_filter else 'All'
    filename   = f'LegitOrganic_Orders_{period}_{status_str}_{date_str}_{time_str}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
