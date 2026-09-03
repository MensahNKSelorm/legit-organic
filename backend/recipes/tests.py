import tempfile
from pathlib import Path

from django.core.management import call_command
from django.core.management.base import CommandError
from django.core.exceptions import ValidationError
from django.test import TestCase, override_settings
from django.utils import timezone
from rest_framework.test import APIClient
from products.models import Product
from decimal import Decimal
from unittest.mock import Mock, patch
from openpyxl import Workbook

from .models import (
    IngredientAlias,
    IngredientMeasurementConversion,
    IngredientNutritionProfile,
    NutritionSourceDataset,
    NutritionSourceRecord,
    Recipe,
    RecipeCombinationNote,
    RecipeIngredient,
    RecipeStep,
    RecipeIngredientProductMatch,
    RegionalNutritionCandidate,
    RecipeSource,
)
from .importing import RecipeImportError, extract_recipe_json_ld, source_for_url, validate_public_url
from .services import (
    calculate_nutrition,
    confirm_regional_candidate,
    normalize_ingredient,
    parse_quantity,
    search_regional_candidates,
    search_usda_candidates,
)
from .wafct import DATA_SHEET, iter_wafct_rows


class DefaultRecipeSearchTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.fufu = Recipe.objects.create(
            title='Fufu', is_default=True, status='approved', is_published=True
        )
        self.light_soup = Recipe.objects.create(
            title='Light Soup', is_default=True, status='approved', is_published=True
        )
        self.groundnut = Recipe.objects.create(
            title='Groundnut Soup',
            description='A rich peanut-based soup',
            is_default=True,
            status='approved',
            is_published=True,
        )
        self.private_recipe = Recipe.objects.create(title='Test Kitchen Draft', is_default=False)
        RecipeIngredient.objects.create(
            recipe=self.light_soup, name='Garden eggs', quantity='4', unit='whole'
        )

    def search(self, query):
        return self.client.get('/api/recipes/default/', {'search': query})

    def test_without_search_returns_only_curated_recipes(self):
        response = self.search('')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 3)

    def test_search_matches_recipe_title(self):
        response = self.search('fufu')
        self.assertEqual([item['title'] for item in response.json()], ['Fufu'])

    def test_plus_search_returns_candidates_for_combined_meal(self):
        response = self.search('fufu + light soup')
        self.assertEqual(
            [item['title'] for item in response.json()],
            ['Fufu', 'Light Soup'],
        )

    def test_search_matches_ingredient_name(self):
        response = self.search('garden eggs')
        self.assertEqual([item['title'] for item in response.json()], ['Light Soup'])

    def test_search_never_exposes_non_curated_recipes(self):
        response = self.search('draft')
        self.assertEqual(response.json(), [])

    def test_public_list_never_exposes_review_drafts(self):
        response = self.client.get('/api/recipes/')
        titles = {item['title'] for item in response.json()}
        self.assertNotIn('Test Kitchen Draft', titles)
        self.assertIn('Fufu', titles)


class RecipeImportSafetyTests(TestCase):
    @patch('recipes.importing._public_addresses', return_value={'93.184.216.34'})
    def test_direct_import_requires_an_enabled_approved_source(self, _addresses):
        RecipeSource.objects.create(
            name='Reviewed publisher',
            base_url='https://recipes.example.com',
            enabled=True,
            allows_recipe_reuse=True,
            reuse_reviewed_at=timezone.now(),
        )
        source = source_for_url(validate_public_url('https://recipes.example.com/tomato-stew'))
        self.assertEqual(source.name, 'Reviewed publisher')
        with self.assertRaises(RecipeImportError):
            source_for_url(validate_public_url('https://unreviewed.example/tomato-stew'))

    @patch('recipes.importing.socket.getaddrinfo')
    def test_private_network_urls_are_rejected(self, getaddrinfo):
        getaddrinfo.return_value = [(None, None, None, None, ('127.0.0.1', 0))]
        with self.assertRaises(RecipeImportError) as error:
            validate_public_url('http://localhost/recipe')
        self.assertEqual(error.exception.code, 'unsafe_url')

    def test_extracts_nested_schema_recipe_before_ai_fallback(self):
        html = '''
        <script type="application/ld+json">
        {"@graph":[{"@type":"Recipe","name":"Tomato stew","recipeYield":"4 servings",
        "prepTime":"PT15M","cookTime":"PT1H","recipeCuisine":"Ghanaian",
        "recipeIngredient":["6 tomatoes","2 onions","2 tbsp oil"],
        "recipeInstructions":[{"@type":"HowToStep","text":"Chop the vegetables."},
        {"@type":"HowToSection","name":"Cook","itemListElement":[{"@type":"HowToStep","text":"Simmer the stew."}]}]}]}
        </script>'''
        result = extract_recipe_json_ld(html)
        self.assertEqual(result['title'], 'Tomato stew')
        self.assertEqual(result['servings'], 4)
        self.assertEqual(result['cook_time'], 60)
        self.assertEqual(result['steps'][1]['section'], 'Cook')


class RecipeShoppingMatchTests(TestCase):
    def test_only_manual_or_verified_matches_reach_customers(self):
        client = APIClient()
        recipe = Recipe.objects.create(title='Tomato stew', status='approved', is_published=True)
        ingredient = RecipeIngredient.objects.create(
            recipe=recipe, name='Tomato', quantity='4', unit='piece'
        )
        suggested = Product.objects.create(name='Tomatoes', price='10.00', unit='basket')
        verified = Product.objects.create(name='Fresh tomatoes', price='12.00', unit='basket')
        RecipeIngredientProductMatch.objects.create(
            recipe_ingredient=ingredient,
            product=suggested,
            match_type='alias',
            confidence='0.750',
            manually_verified=False,
        )
        RecipeIngredientProductMatch.objects.create(
            recipe_ingredient=ingredient,
            product=verified,
            match_type='manual',
            confidence='1.000',
            manually_verified=True,
        )
        response = client.get(f'/api/recipes/{recipe.slug}/')
        matches = response.json()['ingredients'][0]['matched_products']
        self.assertEqual([item['id'] for item in matches], [verified.id])


class RecipePublicationTests(TestCase):
    def test_unreviewed_recipe_cannot_be_published(self):
        recipe = Recipe.objects.create(title='Unreviewed', is_published=True)
        self.assertFalse(recipe.is_published)
        self.assertEqual(recipe.status, 'needs_review')

    def test_rejected_recipe_stays_private(self):
        recipe = Recipe.objects.create(title='Rejected', status='rejected', is_published=True)
        self.assertFalse(recipe.is_published)

    def test_approved_recipe_can_be_published(self):
        recipe = Recipe.objects.create(title='Approved', status='approved', is_published=True)
        self.assertTrue(recipe.is_published)
        self.assertEqual(recipe.status, 'published')


class RecipeCombinationNoteTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        Recipe.objects.create(title='Fufu', is_default=True)
        Recipe.objects.create(title='Light Soup', is_default=True)
        Recipe.objects.create(title='Secret Draft', is_default=False)

    @patch(
        'recipes.views._groq_note',
        return_value=(
            'Peppery light soup lifts the mellow, springy fufu for a balanced spoonful.',
            'test-model',
        ),
    )
    def test_generates_and_caches_note_for_known_dishes(self, generate):
        payload = {'titles': ['Fufu', 'Light Soup']}
        first = self.client.post('/api/recipes/combination-note/', payload, format='json')
        second = self.client.post(
            '/api/recipes/combination-note/', {'titles': ['Light Soup', 'Fufu']}, format='json'
        )

        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.json()['source'], 'generated')
        self.assertEqual(second.json()['source'], 'cache')
        self.assertEqual(RecipeCombinationNote.objects.count(), 1)
        generate.assert_called_once()

    @patch('recipes.views._groq_note', return_value=(None, 'test-model'))
    def test_returns_fallback_without_caching_when_provider_fails(self, _generate):
        response = self.client.post(
            '/api/recipes/combination-note/',
            {'titles': ['Fufu', 'Light Soup']},
            format='json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['source'], 'fallback')
        self.assertEqual(RecipeCombinationNote.objects.count(), 0)

    def test_rejects_unknown_private_and_duplicate_dishes(self):
        for titles in (
            ['Fufu', 'Made-up soup'],
            ['Fufu', 'Secret Draft'],
            ['Fufu', 'fufu'],
        ):
            response = self.client.post(
                '/api/recipes/combination-note/', {'titles': titles}, format='json'
            )
            self.assertEqual(response.status_code, 400)

    def test_requires_between_two_and_four_dishes(self):
        for titles in (
            ['Fufu'],
            ['Fufu', 'Light Soup', 'Groundnut soup', 'Plain rice', 'Kontomire stew'],
        ):
            response = self.client.post(
                '/api/recipes/combination-note/', {'titles': titles}, format='json'
            )
            self.assertEqual(response.status_code, 400)


class IngredientNormalisationTests(TestCase):
    def test_parses_fractions_and_ranges(self):
        self.assertEqual(parse_quantity('1 1/2'), (Decimal('1.5'), None))
        self.assertEqual(parse_quantity('½'), (Decimal('0.5'), None))
        self.assertEqual(parse_quantity('2-3'), (Decimal('2'), Decimal('3')))

    def test_reuses_verified_local_alias_profile(self):
        IngredientAlias.objects.create(alias='kontomire', canonical_name='cocoyam leaves')
        profile = IngredientNutritionProfile.objects.create(
            ingredient_name='Cocoyam leaves',
            normalized_name='cocoyam leaves',
            source='ghanaian_food_composition',
            source_reference='Test source',
            verified=True,
        )
        recipe = Recipe.objects.create(title='Palava sauce')
        ingredient = RecipeIngredient.objects.create(
            recipe=recipe, name='Kontomire', quantity='2', unit='cups'
        )
        normalize_ingredient(ingredient)
        ingredient.refresh_from_db()
        self.assertEqual(ingredient.nutrition_profile, profile)
        self.assertEqual(ingredient.nutrition_match_status, 'local_verified')
        self.assertEqual(ingredient.normalized_unit, 'cup')


class NutritionCalculationTests(TestCase):
    def setUp(self):
        self.recipe = Recipe.objects.create(title='Rice test', servings=2)
        RecipeStep.objects.create(recipe=self.recipe, step_number=1, instruction='Cook.')
        self.profile = IngredientNutritionProfile.objects.create(
            ingredient_name='Rice',
            normalized_name='rice',
            source='usda',
            source_reference='USDA FDC test',
            fdc_id=123,
            verified=True,
            calories_per_100g=Decimal('360'),
            protein_g_per_100g=Decimal('7'),
        )

    def test_calculates_whole_recipe_and_per_serving(self):
        RecipeIngredient.objects.create(
            recipe=self.recipe,
            name='Rice',
            quantity='200',
            unit='g',
            nutrition_profile=self.profile,
        )
        result = calculate_nutrition(self.recipe)
        self.assertTrue(result.is_complete)
        self.assertEqual(result.calories, Decimal('360'))
        self.assertEqual(Decimal(result.total_recipe_values_json['calories']), Decimal('720'))

    def test_unknown_conversion_is_incomplete_not_guessed(self):
        RecipeIngredient.objects.create(
            recipe=self.recipe,
            name='Rice',
            quantity='1',
            unit='cup',
            nutrition_profile=self.profile,
        )
        result = calculate_nutrition(self.recipe)
        self.assertFalse(result.is_complete)
        self.assertIn('no verified cup-to-gram conversion', result.calculation_warnings[0])

    def test_verified_ingredient_specific_conversion_is_used(self):
        IngredientMeasurementConversion.objects.create(
            profile=self.profile,
            unit='cup',
            quantity=1,
            grams=180,
            source_reference='Measured test portion',
            confidence=1,
            verified=True,
        )
        RecipeIngredient.objects.create(
            recipe=self.recipe,
            name='Rice',
            quantity='1',
            unit='cup',
            nutrition_profile=self.profile,
        )
        result = calculate_nutrition(self.recipe)
        self.assertTrue(result.is_complete)
        self.assertEqual(result.calories, Decimal('324'))

    def test_unchanged_fingerprint_reuses_calculation(self):
        RecipeIngredient.objects.create(
            recipe=self.recipe,
            name='Rice',
            quantity='100',
            unit='g',
            nutrition_profile=self.profile,
        )
        first = calculate_nutrition(self.recipe)
        second = calculate_nutrition(self.recipe)
        self.assertEqual(first.pk, second.pk)

    def test_ingredient_change_marks_nutrition_stale(self):
        ingredient = RecipeIngredient.objects.create(
            recipe=self.recipe,
            name='Rice',
            quantity='100',
            unit='g',
            nutrition_profile=self.profile,
        )
        calculate_nutrition(self.recipe)
        ingredient.quantity = '200'
        ingredient.save()
        self.recipe.refresh_from_db()
        self.assertEqual(self.recipe.nutrition_status, 'stale')


class USDACandidateTests(TestCase):
    @patch('recipes.services.requests.post')
    def test_search_saves_candidates_without_silently_selecting_one(self, post):
        response = Mock()
        response.json.return_value = {
            'foods': [
                {
                    'fdcId': 99,
                    'description': 'Peas, black-eyed, dry',
                    'dataType': 'Foundation',
                    'score': 98.2,
                }
            ]
        }
        response.raise_for_status.return_value = None
        post.return_value = response
        recipe = Recipe.objects.create(title='Beans')
        ingredient = RecipeIngredient.objects.create(
            recipe=recipe, name='Black-eyed peas', quantity='100', unit='g'
        )
        with self.settings(USDA_FDC_API_KEY='test-key'):
            candidates = search_usda_candidates(ingredient)
        self.assertEqual(len(candidates), 1)
        ingredient.refresh_from_db()
        self.assertIsNone(ingredient.nutrition_profile)


def make_wafct_fixture():
    workbook = Workbook()
    introduction = workbook.active
    introduction.title = '01 Introduction'
    components = workbook.create_sheet('02 Components')
    components.append(
        [
            'Component in English',
            'Component in French',
            'INFOODS tagname',
            'Unit',
            'Denominator',
            'Significant figures',
            'Maximal decimal places',
            'Datasheet',
            'Analytical/determination method/definition in English',
        ]
    )
    components.append(['Composant', '', 'INFOODS tagname', 'Unité'])
    components.append(['Energy', '', 'ENERC', 'kcal', '/100g EP', 3, 0, '03-06', 'Calculated'])
    components.append(['Protein, total', '', 'PROTCNT', 'g', '/100g EP', 3, 1, '03-06', 'Measured'])
    sheet = workbook.create_sheet(DATA_SHEET)
    headers = [
        'Code',
        'Food name in English',
        'Food name in French',
        'Scientific name',
        'BiblioID/Source',
    ]
    headers.extend(['Energy\n(kcal)', 'Protein, total\n(g)'])
    sheet.append(headers)
    sheet.append(['Code', 'Nom', 'Nom FR', 'Nom scientifique', 'Source'])
    sheet.append(['', '', '', '', '', 'ENERC ', 'PROTCNT '])
    sheet.append(['Roots and tubers'])
    sheet.append(
        [
            '02_039',
            'Cassava, grated, from fermented white cassava, toasted without oil (white gari)',
            'Gari blanc',
            'Manihot esculenta',
            '2P(12)',
            357,
            '[1.2]',
        ]
    )
    handle = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    handle.close()
    workbook.save(handle.name)
    return Path(handle.name)


class WAFCTParserTests(TestCase):
    def setUp(self):
        self.path = make_wafct_fixture()

    def tearDown(self):
        self.path.unlink(missing_ok=True)

    def test_parser_preserves_identity_provenance_and_quality_markers(self):
        row = list(iter_wafct_rows(self.path))[0]
        self.assertEqual(row.food_code, '02_039')
        self.assertIn('white gari', row.original_food_name)
        self.assertIn('fermented', row.preparation_state)
        self.assertIn('toasted', row.preparation_state)
        self.assertEqual(row.source_identifiers['biblio_id_source'], '2P(12)')
        self.assertEqual(row.nutrient_values['ENERC:kcal']['value'], '357')
        self.assertEqual(row.nutrient_values['PROTCNT:g']['value'], '1.2')
        self.assertEqual(row.quality_indicators['PROTCNT:g']['marker'], 'bracketed')

    @override_settings(DEBUG=True)
    def test_local_import_keeps_records_unverified(self):
        call_command('import_wafct', str(self.path))
        record = NutritionSourceRecord.objects.get(food_code='02_039')
        self.assertEqual(record.status, 'unverified')
        self.assertIsNone(record.nutrition_profile)
        self.assertEqual(record.dataset.commercial_permission_status, 'pending')

    @override_settings(DEBUG=False)
    def test_production_import_requires_recorded_commercial_permission(self):
        with self.assertRaisesMessage(CommandError, 'commercial-use permission'):
            call_command('import_wafct', str(self.path))


class RegionalNutritionReviewTests(TestCase):
    def setUp(self):
        self.dataset = NutritionSourceDataset.objects.create(
            code='wafct-2019',
            name='WAFCT',
            version='2019',
            publisher='FAO',
            source_url='https://www.fao.org/',
            citation='WAFCT citation',
            reuse_terms='Commercial permission required.',
            commercial_permission_status='granted',
            commercial_permission_reference='FAO approval ref',
            workbook_sha256='a' * 64,
        )
        self.record = NutritionSourceRecord.objects.create(
            dataset=self.dataset,
            food_code='02_039',
            original_food_name='Cassava, fermented and toasted (white gari)',
            canonical_name='gari',
            preparation_state='fermented, toasted',
            source_identifiers={'biblio_id_source': '2P(12)'},
            nutrient_values={
                'ENERC:kcal': {
                    'tag': 'ENERC',
                    'unit': 'kcal',
                    'denominator': '/100g EP',
                    'value': '357',
                },
                'PROTCNT:g': {
                    'tag': 'PROTCNT',
                    'unit': 'g',
                    'denominator': '/100g EP',
                    'value': '1.2',
                },
            },
            quality_indicators={},
            source_sheet=DATA_SHEET,
            source_row=5,
        )
        self.recipe = Recipe.objects.create(title='Gari test')
        self.ingredient = RecipeIngredient.objects.create(
            recipe=self.recipe,
            name='Gari',
            quantity='100',
            unit='g',
            normalized_ingredient_name='gari',
        )

    @override_settings(DEBUG=False)
    def test_human_confirmation_creates_verified_profile_with_provenance(self):
        candidate = RegionalNutritionCandidate.objects.create(
            recipe_ingredient=self.ingredient,
            source_record=self.record,
        )
        profile = confirm_regional_candidate(candidate, None)
        self.assertTrue(profile.verified)
        self.assertEqual(profile.source, 'wafct_2019')
        self.assertEqual(profile.calories_per_100g, Decimal('357'))
        self.assertEqual(profile.source_metadata['food_code'], '02_039')
        self.record.refresh_from_db()
        self.assertEqual(self.record.status, 'verified')

    def test_source_priority_prefers_manual_then_wafct_then_usda(self):
        IngredientNutritionProfile.objects.create(
            ingredient_name='USDA gari',
            normalized_name='gari',
            source='usda',
            source_reference='USDA',
            verified=True,
        )
        wafct = IngredientNutritionProfile.objects.create(
            ingredient_name='WAFCT gari',
            normalized_name='gari',
            source='wafct_2019',
            source_reference='WAFCT',
            verified=True,
        )
        normalize_ingredient(self.ingredient)
        self.ingredient.refresh_from_db()
        self.assertEqual(self.ingredient.nutrition_profile, wafct)
        manual = IngredientNutritionProfile.objects.create(
            ingredient_name='Verified local gari',
            normalized_name='gari',
            source='manual_verified',
            source_reference='Local laboratory',
            verified=True,
        )
        self.ingredient.nutrition_profile = None
        self.ingredient.save(update_fields=['nutrition_profile'])
        normalize_ingredient(self.ingredient)
        self.ingredient.refresh_from_db()
        self.assertEqual(self.ingredient.nutrition_profile, manual)

    def test_alias_search_creates_candidates_but_never_selects_them(self):
        IngredientAlias.objects.create(
            alias='dawadawa',
            canonical_name='fermented African locust bean',
            lookup_name='fermented African locust bean',
        )
        self.ingredient.name = 'Dawadawa'
        self.ingredient.normalized_ingredient_name = 'fermented african locust bean'
        self.ingredient.save(update_fields=['name', 'normalized_ingredient_name'])
        self.record.original_food_name = 'African locust bean, fermented (soumbala)'
        self.record.save(update_fields=['original_food_name'])
        candidates = search_regional_candidates(self.ingredient)
        self.assertEqual([candidate.source_record for candidate in candidates], [self.record])
        self.ingredient.refresh_from_db()
        self.assertIsNone(self.ingredient.nutrition_profile)

    def test_permission_grant_requires_a_reference(self):
        self.dataset.commercial_permission_status = 'granted'
        self.dataset.commercial_permission_reference = ''
        with self.assertRaises(ValidationError):
            self.dataset.full_clean()
