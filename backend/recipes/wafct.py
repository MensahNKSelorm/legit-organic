import hashlib
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


WAFCT_DATASET_CODE = 'wafct-2019'
WAFCT_DATASET_NAME = 'FAO/INFOODS Food Composition Table for Western Africa'
WAFCT_VERSION = '2019'
WAFCT_SOURCE_URL = 'https://www.fao.org/fileadmin/user_upload/faoweb/2020/WAFCT_2019.xlsx'
WAFCT_CITATION = (
    'Vincent, A. et al. 2020. FAO/INFOODS Food Composition Table for Western '
    'Africa (2019). Rome, FAO.'
)
WAFCT_REUSE_TERMS = (
    'Non-commercial reuse is permitted with acknowledgement. Translation, adaptation, '
    'resale and other commercial use require permission from FAO.'
)
DATA_SHEET = '05 NV_sum_57 (per 100g EP)'
COMPONENT_SHEET = '02 Components'

CORE_NUTRIENT_MAP = {
    ('ENERC', 'kcal'): 'calories_per_100g',
    ('PROTCNT', 'g'): 'protein_g_per_100g',
    ('CHOAVLDF', 'g'): 'carbohydrate_g_per_100g',
    ('FAT', 'g'): 'fat_g_per_100g',
    ('FATCE', 'g'): 'fat_g_per_100g',
    ('FASAT', 'g'): 'saturated_fat_g_per_100g',
    ('FIBTG', 'g'): 'fibre_g_per_100g',
    ('FIBC', 'g'): 'fibre_g_per_100g',
    ('NA', 'mg'): 'sodium_mg_per_100g',
    ('CHOLE', 'mg'): 'cholesterol_mg_per_100g',
}

PREPARATION_TERMS = (
    'raw',
    'boiled',
    'cooked',
    'fried',
    'roasted',
    'steamed',
    'dried',
    'dry',
    'fermented',
    'smoked',
    'grilled',
    'baked',
    'toasted',
    'peeled',
    'unpeeled',
    'with salt',
    'without salt',
    'drained',
    'ripe',
    'unripe',
    'fresh',
    'powder',
    'flour',
)


class WAFCTWorkbookError(ValueError):
    pass


@dataclass(frozen=True)
class WAFCTFoodRow:
    food_code: str
    original_food_name: str
    food_name_french: str
    scientific_name: str
    preparation_state: str
    source_identifiers: dict
    nutrient_values: dict
    quality_indicators: dict
    source_sheet: str
    source_row: int


def workbook_sha256(path):
    digest = hashlib.sha256()
    with Path(path).open('rb') as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def _clean(value):
    return str(value).strip() if value is not None else ''


def _parse_number(raw):
    if raw is None or raw == '':
        return None, None
    if isinstance(raw, (int, float, Decimal)):
        return str(Decimal(str(raw))), None
    text = _clean(raw)
    bracketed = text.startswith('[') and text.endswith(']')
    candidate = text[1:-1].strip() if bracketed else text
    try:
        value = str(Decimal(candidate))
    except InvalidOperation:
        return None, {'raw': text, 'marker': 'non_numeric'}
    return value, ({'raw': text, 'marker': 'bracketed'} if bracketed else None)


def extract_preparation_state(food_name):
    lower = food_name.casefold()
    return ', '.join(term for term in PREPARATION_TERMS if term in lower)


def inspect_components(workbook):
    if COMPONENT_SHEET not in workbook.sheetnames:
        raise WAFCTWorkbookError(f'Missing required sheet: {COMPONENT_SHEET}')
    components = {}
    sheet = workbook[COMPONENT_SHEET]
    for row in sheet.iter_rows(min_row=3, values_only=True):
        tag = _clean(row[2]).replace(' ', '')
        unit = _clean(row[3])
        if not tag:
            continue
        for individual_tag in re.split(r'\s+or\s+|\s+OR\s+', tag):
            individual_tag = individual_tag.strip('[] ')
            if individual_tag:
                components[(individual_tag, unit)] = {
                    'name': _clean(row[0]),
                    'unit': unit,
                    'denominator': _clean(row[4]),
                    'definition': _clean(row[8]),
                }
    return components


def iter_wafct_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if DATA_SHEET not in workbook.sheetnames:
        raise WAFCTWorkbookError(f'Missing required sheet: {DATA_SHEET}')
    components = inspect_components(workbook)
    sheet = workbook[DATA_SHEET]
    headers = [_clean(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    tags = [
        _clean(sheet.cell(3, col).value).replace(' ', '') for col in range(1, sheet.max_column + 1)
    ]
    for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), 5):
        code = _clean(row[0])
        if not re.fullmatch(r'\d{2}_\d{3}', code):
            continue
        food_name = _clean(row[1])
        nutrients, quality = {}, {}
        for index in range(5, len(row)):
            tag = tags[index]
            if not tag:
                continue
            unit_match = re.search(r'\(([^)]+)\)', headers[index])
            unit = unit_match.group(1) if unit_match else '-'
            value, marker = _parse_number(row[index])
            key = f'{tag}:{unit}'
            nutrients[key] = {
                'tag': tag,
                'unit': unit,
                'denominator': '/100g EP',
                'value': value,
                'raw_value': _clean(row[index]),
                'component': components.get((tag, unit), {}).get('name', headers[index]),
            }
            if marker:
                quality[key] = marker
        yield WAFCTFoodRow(
            food_code=code,
            original_food_name=food_name,
            food_name_french=_clean(row[2]),
            scientific_name=_clean(row[3]),
            preparation_state=extract_preparation_state(food_name),
            source_identifiers={'biblio_id_source': _clean(row[4])},
            nutrient_values=nutrients,
            quality_indicators=quality,
            source_sheet=DATA_SHEET,
            source_row=row_number,
        )


def profile_values_from_record(record):
    values = {
        'calories_per_100g': None,
        'protein_g_per_100g': None,
        'carbohydrate_g_per_100g': None,
        'fat_g_per_100g': None,
        'saturated_fat_g_per_100g': None,
        'fibre_g_per_100g': None,
        'sugar_g_per_100g': None,
        'sodium_mg_per_100g': None,
        'cholesterol_mg_per_100g': None,
    }
    micronutrients = {}
    for nutrient in record.nutrient_values.values():
        raw_value = nutrient.get('value')
        if raw_value is None:
            continue
        target = CORE_NUTRIENT_MAP.get((nutrient['tag'], nutrient['unit']))
        if target:
            values[target] = Decimal(raw_value)
        else:
            micronutrients[nutrient['tag']] = {
                'value': raw_value,
                'unit': nutrient['unit'],
                'denominator': nutrient['denominator'],
            }
    values['micronutrients_json'] = micronutrients
    return values
