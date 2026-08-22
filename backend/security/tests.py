from io import BytesIO, StringIO
from unittest.mock import patch

from django.contrib import admin
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import RequestFactory, TestCase
from django.test import modify_settings, override_settings
from django.urls import reverse
from django_otp.oath import totp
from django_otp.plugins.otp_totp.models import TOTPDevice
from openpyxl import load_workbook
from axes.utils import reset as reset_axes

from users.models import User
from blog.models import BlogPost
from orders.models import Order
from products.models import Product

from .audit import record_event
from .admin import AuditEventAdmin, export_audit_events
from .models import AuditEvent, StaffSecurityProfile


class AuditEventTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_superuser(
            email='owner@example.com', password='x', first_name='Owner', last_name='User'
        )
        self.request = RequestFactory().post(
            '/admin/orders/order/1/change/', HTTP_USER_AGENT='Test Browser',
            HTTP_X_FORWARDED_FOR='203.0.113.9, 127.0.0.1',
        )
        self.request.user = self.owner

    def test_records_actor_target_network_and_redacts_secrets(self):
        event = record_event(
            action='payment.corrected', request=self.request, target=self.owner,
            reason='Verified manual payment', before={'status': 'pending', 'token': 'secret'},
            after={'status': 'success', 'password': 'never-store-this'},
        )
        self.assertEqual(event.actor, self.owner)
        self.assertEqual(event.ip_address, '203.0.113.9')
        self.assertEqual(event.before['token'], '[REDACTED]')
        self.assertEqual(event.after['password'], '[REDACTED]')

    def test_event_cannot_be_changed_or_deleted(self):
        event = record_event(action='security.test', actor=self.owner)
        event.reason = 'tampered'
        with self.assertRaises(RuntimeError):
            event.save()
        with self.assertRaises(RuntimeError):
            event.delete()


class StaffTwoFactorTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            email='staff@legitorganic.com', password='StrongPass123!',
            first_name='Staff', last_name='Member', is_staff=True,
        )
        self.client.force_login(self.staff)

    def test_enrollment_confirms_device_and_returns_recovery_codes_once(self):
        url = reverse('staff-security:setup')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Connect your authenticator')
        device = TOTPDevice.objects.get(user=self.staff, confirmed=False)
        token = str(totp(
            device.bin_key, step=device.step, t0=device.t0,
            digits=device.digits, drift=device.drift,
        )).zfill(device.digits)
        response = self.client.post(url, {
            'current_password': 'StrongPass123!', 'token': token,
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'RECOVERY CODES')
        device.refresh_from_db()
        self.assertTrue(device.confirmed)
        self.assertEqual(self.staff.staff_security.recovery_codes.count(), 10)
        self.assertTrue(AuditEvent.objects.filter(action='security.2fa_enrolled').exists())

    @override_settings(STAFF_2FA_MODE='enforce')
    @modify_settings(MIDDLEWARE={'append': 'security.middleware.StaffSecurityMiddleware'})
    def test_enforcement_redirects_unenrolled_staff_to_setup(self):
        response = self.client.get('/admin/')
        self.assertRedirects(response, reverse('staff-security:setup'))

    @override_settings(STAFF_2FA_MODE='enroll', STAFF_OWNER_2FA_REQUIRED=True)
    @modify_settings(MIDDLEWARE={'append': 'security.middleware.StaffSecurityMiddleware'})
    def test_enrollment_phase_requires_owner_to_enroll_first(self):
        owner = User.objects.create_superuser(
            email='owner@legitorganic.com', password='OwnerPass123!',
            first_name='Owner', last_name='User',
        )
        self.client.force_login(owner)
        response = self.client.get('/admin/')
        self.assertRedirects(response, reverse('staff-security:setup'))

    @override_settings(
        STAFF_2FA_MODE='enroll', STAFF_IDLE_TIMEOUT_SECONDS=30,
        STAFF_ABSOLUTE_SESSION_SECONDS=3600,
    )
    @modify_settings(MIDDLEWARE={'append': 'security.middleware.StaffSecurityMiddleware'})
    def test_idle_staff_session_expires(self):
        session = self.client.session
        session['staff_session_started_at'] = 100
        session['staff_session_last_seen_at'] = 100
        session.save()
        response = self.client.get('/admin/')
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('admin:login'), response.url)
        self.assertTrue(AuditEvent.objects.filter(action='security.session_expired').exists())

    def test_verify_rejects_external_next_redirect(self):
        device = TOTPDevice.objects.create(
            user=self.staff, name='Authenticator', confirmed=True,
        )
        token = str(totp(
            device.bin_key, step=device.step, t0=device.t0,
            digits=device.digits, drift=device.drift,
        )).zfill(device.digits)
        response = self.client.post(
            reverse('staff-security:verify') + '?next=https://evil.example/steal',
            {'token': token},
        )
        self.assertRedirects(response, '/admin/', fetch_redirect_response=False)


class StaffSecurityResetTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_superuser(
            email='owner@legitorganic.com', password='OwnerPass123!',
            first_name='Owner', last_name='User',
        )
        self.owner_device = TOTPDevice.objects.create(
            user=self.owner, name='Owner authenticator', confirmed=True,
        )
        self.staff = User.objects.create_user(
            email='staff@legitorganic.com', password='StaffPass123!',
            first_name='Staff', last_name='Member', is_staff=True,
        )
        TOTPDevice.objects.create(
            user=self.staff, name='Staff authenticator', confirmed=True,
        )
        StaffSecurityProfile.objects.create(user=self.staff, enrolled_at=self.owner.date_joined)
        self.client.force_login(self.owner)

    def _owner_token(self):
        device = self.owner_device
        return str(totp(
            device.bin_key, step=device.step, t0=device.t0,
            digits=device.digits, drift=device.drift,
        )).zfill(device.digits)

    @patch('security.views.verify_staff_code', return_value=(True, False))
    def test_owner_reset_requires_reason_password_and_2fa_then_audits(self, _verify):
        url = reverse('staff-security:reset-staff', args=[self.staff.pk])
        denied = self.client.post(url, {
            'reason': '', 'current_password': 'wrong', 'otp_token': '000000',
        })
        self.assertEqual(denied.status_code, 200)
        self.assertTrue(TOTPDevice.objects.filter(user=self.staff, confirmed=True).exists())

        allowed = self.client.post(url, {
            'reason': 'Staff replaced their phone',
            'current_password': 'OwnerPass123!',
            'otp_token': '123456',
        })
        self.assertEqual(allowed.status_code, 200)
        self.assertFalse(TOTPDevice.objects.filter(user=self.staff).exists())
        event = AuditEvent.objects.get(action='security.2fa_reset')
        self.assertEqual(event.actor, self.owner)
        self.assertEqual(event.target_id, str(self.staff.pk))
        self.assertEqual(event.reason, 'Staff replaced their phone')

    def test_non_owner_cannot_reset_staff_2fa(self):
        self.client.force_login(self.staff)
        response = self.client.get(
            reverse('staff-security:reset-staff', args=[self.staff.pk])
        )
        self.assertEqual(response.status_code, 403)


class AuditAdminTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_superuser(
            email='owner@example.com', password='x', first_name='Owner', last_name='User'
        )
        self.staff = User.objects.create_user(
            email='staff@example.com', password='x', first_name='Staff',
            last_name='User', is_staff=True,
        )
        self.owner_event = record_event(action='owner.action', actor=self.owner)
        self.staff_event = record_event(action='staff.action', actor=self.staff)

    def test_regular_staff_only_sees_own_events_and_cannot_export(self):
        request = RequestFactory().get('/admin/security/auditevent/')
        request.user = self.staff
        model_admin = AuditEventAdmin(AuditEvent, admin.site)
        self.assertEqual(list(model_admin.get_queryset(request)), [self.staff_event])
        self.assertEqual(model_admin.get_actions(request), {})

    def test_owner_export_is_an_excel_workbook(self):
        request = RequestFactory().post('/admin/security/auditevent/')
        request.user = self.owner
        response = export_audit_events(
            None, request, AuditEvent.objects.order_by('created_at')
        )
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook['Security Audit']
        self.assertEqual(sheet['A1'].value, 'Timestamp')
        self.assertEqual(sheet.max_row, 4)
        self.assertTrue(
            AuditEvent.objects.filter(action='security.audit_exported').exists()
        )


class StaffLoginLockoutTests(TestCase):
    def setUp(self):
        reset_axes()
        self.staff = User.objects.create_user(
            email='locked@legitorganic.com', password='CorrectPass123!',
            first_name='Locked', last_name='Staff', is_staff=True,
        )

    def tearDown(self):
        reset_axes()

    def test_five_failures_lock_staff_account_and_are_audited(self):
        url = reverse('admin:login')
        for _ in range(5):
            response = self.client.post(
                url,
                {'username': self.staff.email, 'password': 'WrongPass123!'},
                REMOTE_ADDR='203.0.113.44',
            )
        self.assertIn(response.status_code, (200, 403, 429))
        locked = self.client.post(
            url,
            {'username': self.staff.email, 'password': 'CorrectPass123!'},
            REMOTE_ADDR='203.0.113.44',
        )
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertTrue(
            AuditEvent.objects.filter(action='security.login_locked').exists()
        )


class StaffTwoFactorReadinessCommandTests(TestCase):
    def test_command_blocks_enforcement_until_every_active_staff_is_enrolled(self):
        owner = User.objects.create_superuser(
            email='owner@legitorganic.com', password='x',
            first_name='Owner', last_name='User',
        )
        with self.assertRaises(CommandError):
            call_command('check_staff_2fa_readiness', stdout=StringIO())
        TOTPDevice.objects.create(
            user=owner, name='Owner authenticator', confirmed=True,
        )
        output = StringIO()
        call_command('check_staff_2fa_readiness', stdout=output)
        self.assertIn('safe to enable', output.getvalue())


class ExceptionalDataActionTests(TestCase):
    def setUp(self):
        self.owner = User.objects.create_superuser(
            email='owner@legitorganic.com', password='OwnerPass123!',
            first_name='Owner', last_name='User',
        )
        self.customer = User.objects.create_user(
            email='customer@example.com', password='CustomerPass123!',
            first_name='Ama', last_name='Mensah', phone_number='0244000000',
            city='Accra', street_address='Market Road', email_verified=True,
        )
        self.order = Order.objects.create(
            user=self.customer, reference='LO-KEEP-HISTORY', delivery_address='Accra',
            total_amount='25.00', order_source='whatsapp',
        )
        self.client.force_login(self.owner)

    @patch('security.views.verify_staff_code', return_value=(True, False))
    def test_customer_anonymization_removes_identity_but_keeps_order(self, _verify):
        response = self.client.post(
            reverse('staff-security:anonymize-customer', args=[self.customer.pk]),
            {
                'reason': 'Approved privacy request',
                'current_password': 'OwnerPass123!', 'otp_token': '123456',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.customer.refresh_from_db()
        self.order.refresh_from_db()
        self.assertFalse(self.customer.is_active)
        self.assertTrue(self.customer.email.endswith('@anonymized.invalid'))
        self.assertEqual(self.customer.phone_number, '')
        self.assertEqual(self.order.user_id, self.customer.pk)
        self.assertEqual(self.order.delivery_address, '[Anonymized]')
        self.assertTrue(AuditEvent.objects.filter(action='customer.anonymized').exists())

    @patch('security.views.verify_staff_code', return_value=(True, False))
    def test_owner_can_exceptionally_delete_blog_with_audit_reason(self, _verify):
        post = BlogPost.objects.create(title='Remove me', content='<p>Draft</p>')
        response = self.client.post(
            reverse(
                'staff-security:exceptional-delete',
                args=['blog', 'blogpost', post.pk],
            ),
            {
                'reason': 'Duplicate draft approved for permanent deletion',
                'current_password': 'OwnerPass123!', 'otp_token': '123456',
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(BlogPost.objects.filter(pk=post.pk).exists())
        event = AuditEvent.objects.get(action='content.permanently_deleted')
        self.assertEqual(event.reason, 'Duplicate draft approved for permanent deletion')
        self.assertEqual(event.before['target_type'], 'blog.blogpost')

    def test_products_cannot_be_permanently_deleted_because_orders_need_history(self):
        product = Product.objects.create(name='Historical product', price='10.00')
        response = self.client.get(reverse(
            'staff-security:exceptional-delete',
            args=['products', 'product', product.pk],
        ))
        self.assertEqual(response.status_code, 403)
        self.assertTrue(Product.objects.filter(pk=product.pk).exists())
