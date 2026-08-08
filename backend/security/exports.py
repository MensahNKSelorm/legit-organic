import json

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def generate_audit_excel(events):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Security Audit'
    headers = [
        'Timestamp', 'Severity', 'Action', 'Staff', 'Target type', 'Target',
        'Reason', 'Before', 'After', 'IP address', 'User agent',
    ]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=column, value=value)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0D3B2A')

    for row, event in enumerate(events, 2):
        values = [
            timezone.localtime(event.created_at).strftime('%Y-%m-%d %H:%M:%S'),
            event.get_severity_display(), event.action, event.actor_email,
            event.target_type, event.target_label, event.reason,
            json.dumps(event.before, ensure_ascii=False),
            json.dumps(event.after, ensure_ascii=False),
            str(event.ip_address or ''), event.user_agent,
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row=row, column=column, value=value)

    widths = [21, 14, 28, 30, 24, 38, 40, 45, 45, 18, 45]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = timezone.now().strftime('LegitOrganic_SecurityAudit_%Y-%m-%d_%H%M.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response
